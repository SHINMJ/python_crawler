import sys
from urllib.parse import quote_plus
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup as bs, element
import time 
from openpyxl import Workbook
from openpyxl import load_workbook
import platform
from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QProgressBar

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()
        

    
    def initUI(self):
        self.setWindowTitle('Naver Map Crawler')
        
        lbl = QLabel(self)
        lbl.move(60,40)
        lbl.setText('검색어 입력 후 엔터')
        
        self.qle = QLineEdit(self)
        self.qle.move(60, 70)
        self.qle.returnPressed.connect(self.on_line_edit_returnPressed)

        self.setGeometry(50, 50, 250, 200)
        self.progressbar = QProgressBar(self)
        self.progressbar.setGeometry(20, 100, 200, 25)
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(4)

    
    def on_line_edit_returnPressed(self):
       print(self.qle.text())
       self.run()


    def getPlatform(self) :
        os = platform.system()
        if os == 'Windows':
            return 'windows'
        elif os == 'Darwin':
            return 'macosm1'
        elif os == 'Linux':
            return 'linux'


    def readData(self, driver, lis, index) :
        datalist = []
        for idx, li in enumerate(lis):
            el = li.find('span', '_3Apve')
            name = el.get_text()
            
            aele = ''
            try:
                aele = driver.find_element_by_css_selector('#_pcmap_list_scroll_container > ul > li:nth-child('+str(index+idx+1)+') > div:nth-child(2) > a:nth-child(1)') 
            except:
                aele = driver.find_element_by_css_selector('#_pcmap_list_scroll_container > ul > li:nth-child('+str(index+idx+1)+') > div:nth-child(1) > a:nth-child(1)')

            # print(name)
            aele.click()
        
            time.sleep(3)

            driver.switch_to.default_content()
            driver.switch_to.frame('entryIframe')

            # print(driver.page_source)
            subhtml = bs(driver.page_source, 'html.parser')
            span = subhtml.select('div.place_section_content > ul > li > div > span')
            phone = span[0].get_text()
            addr = span[2].get_text()
            datalist.append({'name': name, 'phone': phone, 'address': addr})
            driver.switch_to.default_content()
            driver.switch_to.frame('searchIframe')
            time.sleep(1)
        return datalist


    def getData(self, searchstr) :
        # https://map.naver.com/v5/search/경기%20시흥%20왁싱?
        query = quote_plus(searchstr)
        url = f'http://map.naver.com/v5/search/{query}'

        path = self.getPlatform()
        print(path)
        chromedriver = f'./driver/{path}/chromedriver'
        options = webdriver.ChromeOptions()
        options.add_argument('headless')        # headless chrome
        options.add_argument('disable-gpu')    # GPU 사용 안함
        options.add_argument('lang=ko_KR')    # 언어 설정

        driver = webdriver.Chrome(chromedriver, options=options)

        driver.get(url)

        datalist = []

        driver.implicitly_wait(5)

        #loc-main-section-root > section > div > div:nth-child(2) > ul
        #loc-main-section-root > section > div > div:nth-child(2) > ul > li:nth-child(1)

        self.progressbar.setValue(1)

        while True:
            time.sleep(3)
            
            print('while!!!!!!!!!!!')
            driver.switch_to.frame("searchIframe")

            print(driver.page_source)

            time.sleep(2)

            #_pcmap_list_scroll_container > ul > li:nth-child(45) > div._3ZU00._1rBq3 > a:nth-child(1)
            total = 0
            for i in range(0, 5):
                js_script = "document.querySelector(\"#app-root\").innerHTML"
                raw = driver.execute_script("return " + js_script)

                html = bs(raw, "html.parser")

                lis = html.select('#_pcmap_list_scroll_container > ul > li')
                if i == 0: total = len(lis)
                    
                if total == 10:
                    readlis = lis[(10*i):]
                    datalist = datalist + self.readData( driver=driver, lis=readlis, index=(10*i))
                else:
                    datalist = datalist + self.readData( driver=driver, lis=lis, index=(10*i))
                    break
                    

            try:
                nextel = driver.find_element_by_css_selector('div._2ky45 > a:last-child') 
                if '_34lTS' in nextel.get_attribute('class'):
                    break
                
                nextel.click()
                print('click????????????')
                
            except:
                print('done!!!!!')
                break
            finally:
                driver.switch_to.default_content()
                driver.quit()
        
        if driver:
            driver.quit()

        return datalist
        

    def writeToExcel(self, write_wb, datalist, query):

        write_ws = write_wb.create_sheet(query)

        write_ws = write_wb[query]

        # header 설정
        write_ws['A1'] = '상호'
        write_ws['B1'] = '전화번호'
        write_ws['C1'] = '주소'

        self.progressbar.setValue(3)
        for idx, data in enumerate(datalist):
            write_ws.cell(row=(idx+2), column=1).value = data['name']
            write_ws.cell(row=(idx+2), column=2).value = data['phone']
            write_ws.cell(row=(idx+2), column=3).value = data['address']

        return write_wb


    def run(self):
        query = self.qle.text()
        self.progressbar.setValue(0)
        data = self.getData(searchstr=query)
        print(data)
        
        try:
            wb = load_workbook('./data.xlsx')
        except:
            wb = Workbook()

        
        self.progressbar.setValue(2)
        wb = self.writeToExcel(write_wb=wb, datalist=data, query=query)
        wb.save(f'./data.xlsx')
        self.progressbar.setValue(4)

    

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    ex.show()
    sys.exit(app.exec_())





